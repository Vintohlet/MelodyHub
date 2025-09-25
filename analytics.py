import os
import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from sqlalchemy import create_engine

sys.stdout.reconfigure(encoding="utf-8")
load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
engine = create_engine(DATABASE_URL)

output_dir = "graphs"
os.makedirs(output_dir, exist_ok=True)


def load_queries(path: str) -> dict:
    queries = {}
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    blocks = content.split("-- name:")
    for block in blocks:
        if not block.strip():
            continue
        name, sql = block.strip().split("\n", 1)
        queries[name.strip()] = sql.strip().rstrip(";")
    return queries


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip().str.lower()
    return df


def prepare_for_export(df: pd.DataFrame, qname: str) -> pd.DataFrame:
    df = clean_columns(df).copy()

    if qname == "q4":
        if "year" in df.columns and "revenue" in df.columns:
            df = df[["year", "revenue"]]

    if qname == "q6":
        df = df.drop(columns=["customer_id"], errors="ignore")
        
        for col in ["first_purchase", "last_purchase"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d")

    return df


def visualize(df: pd.DataFrame, qname: str):
    df = clean_columns(df)
    fig, ax = plt.subplots(figsize=(10, 8))

    if qname == "q1":
        df.set_index("genre")["tracks_sold"].plot.pie(autopct="%1.1f%%", ax=ax)
        plt.title("Топ-10 жанров по числу проданных треков")

    elif qname == "q2":
        df.set_index("album")["revenue"].plot.barh(ax=ax)
        plt.title("Топ-10 альбомов по выручке")
        plt.xlabel("Выручка ($)")
        plt.ylabel("Альбом")

    elif qname == "q3":
        df.set_index("country")["revenue"].plot.bar(ax=ax)
        plt.title("Выручка по странам")
        plt.xlabel("Страна")
        plt.ylabel("Выручка ($)")
        plt.xticks(rotation=45)

    elif qname == "q4":
        df["year_num"] = df["year"].astype(float).astype(int)
        df["revenue_num"] = df["revenue"].astype(float)

        df = df.drop_duplicates(subset=["year_num"])

        df.plot.line(x="year_num", y="revenue_num", marker="o", ax=ax)
        plt.title("Выручка по годам")
        plt.xlabel("Год")
        plt.ylabel("Выручка ($)")
 
        ax.set_xticks(df["year_num"])
        ax.set_xticklabels(df["year_num"].astype(int))

    elif qname == "q5":
        df["avg_length_min"].plot.hist(bins=10, edgecolor="black", ax=ax)
        plt.title("Распределение средней длины треков по жанрам")
        plt.xlabel("Длина трека (мин)")
        plt.ylabel("Количество жанров")

    elif qname == "q6":
        df["first_purchase"] = pd.to_datetime(df["first_purchase"])
        df["last_purchase"] = pd.to_datetime(df["last_purchase"])
        df["loyalty_days"] = (df["last_purchase"] - df["first_purchase"]).dt.days

        scatter = ax.scatter(
            df["first_purchase"],
            df["revenue"],
            s=df["revenue"] * 10,
            c=df["loyalty_days"],
            cmap="viridis",
            alpha=0.7,
            edgecolors="black"
        )

        for _, row in df.iterrows():
            ax.annotate(
                row["customer"],
                (row["first_purchase"], row["revenue"]),
                fontsize=8,
                xytext=(5, 5),
                textcoords="offset points"
            )

        cbar = plt.colorbar(scatter)
        cbar.set_label("Лояльность (дни между первой и последней покупкой)")
        plt.title("Клиенты: первая покупка vs выручка")
        plt.xlabel("Дата первой покупки")
        plt.ylabel("Выручка ($)")
        plt.grid(True, linestyle="--", alpha=0.6)

    plt.tight_layout()
    plt.savefig(f"{output_dir}/{qname}_chart.png", dpi=300, bbox_inches="tight")
    plt.close()
    print(f"График {qname} сохранен в {output_dir}/{qname}_chart.png")


def create_interactive_time_slider(df: pd.DataFrame, qname: str):
    df = clean_columns(df)
    if qname == "q4":
        df['year_num'] = df['year'].astype(int)
        df['revenue_num'] = df['revenue'].astype(float)

        df_sorted = df.sort_values('year_num')
        df_expanded = []
        for _, row in df_sorted.iterrows():
            for _, target_row in df_sorted.iterrows():
                if target_row['year_num'] <= row['year_num']:
                    df_expanded.append({
                        "year": target_row['year_num'],
                        "revenue": target_row['revenue_num'],
                        "animation_year": row['year_num']
                    })

        df_plot = pd.DataFrame(df_expanded)
        fig = px.bar(
            df_plot,
            x="year",
            y="revenue",
            animation_frame="animation_year",
            title="Выручка по годам - временная динамика",
            color="revenue",
            color_continuous_scale="viridis"
        )

        fig.update_layout(
            width=1000,
            height=600,
            yaxis_range=[0, df['revenue_num'].max() * 1.1],
            xaxis=dict(
                tickmode='array',
                tickvals=sorted(df['year_num'].unique()),
                ticktext=[str(int(year)) for year in sorted(df['year_num'].unique())]
            )
        )
        fig.show()
        return fig
    else:
        return None


def export_to_excel(dataframes_dict: dict, filename: str):
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)

 
    total_rows = 0
    

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df_prepared = prepare_for_export(df, sheet_name)
            df_prepared.to_excel(writer, sheet_name=sheet_name, index=False)
    
    wb = load_workbook(filepath)
    
    for sheet_name, df in dataframes_dict.items():
        ws = wb[sheet_name]
        total_rows += ws.max_row - 1 

       
        ws.freeze_panes = "B2"

        ws.auto_filter.ref = ws.dimensions

        df_prepared = prepare_for_export(df, sheet_name)
        for col_idx, col in enumerate(df_prepared.columns, 1):
           
            if (pd.api.types.is_numeric_dtype(df_prepared[col]) and 
                col.lower() not in ["year", "customer_id"]):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(cell_range, rule)

       
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(filepath)


    sheet_count = len(dataframes_dict)
    if sheet_count == 1:
        sheets_word = "лист"
    elif 2 <= sheet_count <= 4:
        sheets_word = "листа"
    else:
        sheets_word = "листов"
    
    print(f"Создан файл {filename}, {sheet_count} {sheets_word}, {total_rows} строк")


if __name__ == "__main__":
    print("Загружаем и обрабатываем запросы из файла queries.sql...")
    queries = load_queries("queries.sql")
    dataframes_dict = {}

    for qname, query in queries.items():
        print(f"Выполняем запрос {qname}...")
        df = pd.read_sql_query(query, engine)
        dataframes_dict[qname] = df
        visualize(df, qname)

        if qname == "q4":
            print("\nСоздаем интерактивный график с временным ползунком для q4...")
            create_interactive_time_slider(df, qname)

    export_to_excel(dataframes_dict, "music_report.xlsx")

    print(f"\nВсе статические графики сохранены в папку '{output_dir}'")